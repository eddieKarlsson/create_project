import os
import shutil
import sys
import openpyxl as xl

DOC_PATH = r'C:\Users\EddieKarlsson\Documents'
TEMPLATE_PATH = r'project_structure\free'

new_proj_name = sys.argv[1]
dest_path = os.path.join(sys.argv[2], sys.argv[1])
projnr = sys.argv[3]

def _openpyxl_open_workbook_and_edit_attr(excel_path):
    try:
        wb = xl.load_workbook(excel_path)
    except FileNotFoundError as e:
        print(e)
        printf('ERROR! Excel {excel_path} file not found, program will exit')
        sys.exit()

    ws = wb['Overview']
    #ws = wb.active

    # Shift index right in sheet until "Projectnumber" text is found
    for i in range(2, 99):
        cell = ws.cell(row=1, column=i)

        if cell.value and 'number' in cell.value:
            print(cell)
            col_index = i
            break
        else:
            continue

        # Write to cells
        ws.cell(row=1, column=col_index).value = projnr
        ws.cell(row=2, column=col_index).value = new_proj_name

    wb.save(excel_path)



source_path = os.path.join(DOC_PATH, TEMPLATE_PATH)

shutil.copytree(source_path, dest_path)

wb_src_path = os.path.join(DOC_PATH, 'MC Doc\WB', 'latest_wb.xlsx')
wb_dst_path = os.path.join(dest_path, 'doc', new_proj_name + '_wb.xlsx')
shutil.copy(wb_src_path, wb_dst_path)

td_src_path = os.path.join(DOC_PATH, 'MC Doc\TD', 'latest_td.xlsx')
td_dst_path = os.path.join(dest_path, 'doc', new_proj_name + '_td.xlsx')
shutil.copy(td_src_path, td_dst_path)

_open_workbook_and_edit_attr(td_dst_path)
